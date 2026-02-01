import streamlit as st
import pandas as pd
import numpy as np
import random
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font

# --- 頁面設定 ---
st.set_page_config(page_title="自動排班系統", layout="wide")

st.title("📅 員工自動排班管理系統")
st.markdown("### 規則：週一二限5人 / 週三四五限6人 / 優先權與轉讓機制")

# --- 側邊欄：設定區 ---
with st.sidebar:
    st.header("⚙️ 參數設定")
    year = st.number_input("年份", min_value=2024, max_value=2030, value=2026)
    month = st.number_input("月份", min_value=1, max_value=12, value=2)
    
    st.divider()
    
    st.subheader("⭐ 本月優先名單")
    all_employees = [f"a{i}" for i in range(1, 25)]
    priority_emps = st.multiselect(
        "選擇享有優先權的員工 (4位)",
        options=all_employees,
        default=["a1", "a7", "a11", "a14"]
    )
    st.info(f"目前優先人員：{', '.join(priority_emps)}")

# --- 邏輯函數 ---
def generate_template_bytes():
    output = io.BytesIO()
    employees = [f"a{i}" for i in range(1, 25)]
    days = list(range(1, 32))
    df_requests = pd.DataFrame(index=employees, columns=days)
    df_requests.iloc[0,0] = "填1(事)2(公)"
    df_transfers = pd.DataFrame(columns=["Date", "From", "To"])
    df_transfers.loc[0] = [3, "a1", "a2"] 
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_requests.to_excel(writer, sheet_name="Requests")
        df_transfers.to_excel(writer, sheet_name="Transfers", index=False)
    return output.getvalue()

def process_schedule(uploaded_file, year, month, priority_emps):
    try:
        df_requests = pd.read_excel(uploaded_file, sheet_name="Requests", index_col=0)
        df_transfers = pd.read_excel(uploaded_file, sheet_name="Transfers")
    except Exception as e:
        return None, None, f"讀取錯誤：{e}"

    approved_set = set()
    display_df = df_requests.copy()
    display_df[:] = "" 
    days_cols = list(df_requests.columns)
    weekday_map = {0:"一", 1:"二", 2:"三", 3:"四", 4:"五", 5:"六", 6:"日"}
    weekday_row_data = []
    
    for day in days_cols:
        try:
            day_num = int(day)
            date_obj = pd.Timestamp(year=year, month=month, day=day_num)
            wk_idx = date_obj.dayofweek
            weekday_row_data.append(weekday_map[wk_idx])
        except:
            weekday_row_data.append("")
            continue
            
        if wk_idx >= 5: continue 
        limit = 5 if wk_idx <= 1 else 6
        
        day_col = df_requests[day]
        requesting_emps = day_col[day_col.notna() & (day_col != 0)].index.tolist()
        if not requesting_emps: continue
            
        priority_group = []
        official_group = []
        regular_group = []
        daily_transfers = df_transfers[df_transfers["Date"] == day_num]
        transfer_map = {}
        for _, row in daily_transfers.iterrows():
            transfer_map[str(row["From"])] = str(row["To"])
            
        for emp in requesting_emps:
            emp_str = str(emp)
            req_type = day_col[emp]
            is_priority = False
            if emp_str in priority_emps: is_priority = True
            for giver, receiver in transfer_map.items():
                if receiver == emp_str:
                    if giver in priority_emps and giver not in requesting_emps:
                        is_priority = True
                    break
            
            if is_priority: priority_group.append(emp)
            elif req_type == 2: official_group.append(emp)
            else: regular_group.append(emp)
            
        current_approved = []
        current_approved.extend(priority_group)
        current_approved.extend(official_group)
        remaining = limit - len(current_approved)
        if remaining > 0:
            if len(regular_group) <= remaining:
                current_approved.extend(regular_group)
            else:
                winners = random.sample(regular_group, remaining)
                current_approved.extend(winners)
                
        for emp in current_approved:
            approved_set.add((emp, day))
            original_val = df_requests.loc[emp, day]
            txt = "公" if original_val == 2 else "休"
            display_df.loc[emp, day] = txt

    wk_df = pd.DataFrame([weekday_row_data], columns=days_cols, index=["星期"])
    final_display_df = pd.concat([wk_df, display_df])
    return final_display_df, approved_set, None

def create_download_excel(df_requests, approved_set, year, month):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "排班結果"
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    ws.cell(1, 1, "日期"); ws.cell(2, 1, "星期"); ws.cell(3, 1, "員工")
    days_cols = list(df_requests.columns)
    weekday_map = {0:"一", 1:"二", 2:"三", 3:"四", 4:"五", 5:"六", 6:"日"}
    
    for col_idx, day in enumerate(days_cols, start=2):
        ws.cell(1, col_idx, day).alignment = center_align
        try:
            d_obj = pd.Timestamp(year=year, month=month, day=int(day))
            wk = weekday_map[d_obj.dayofweek]
            c = ws.cell(2, col_idx, wk)
            c.alignment = center_align
            if d_obj.dayofweek >= 5: c.font = Font(color="808080")
        except: pass
        
    for row_idx, emp in enumerate(df_requests.index, start=3):
        ws.cell(row_idx, 1, emp)
        for col_idx, day in enumerate(days_cols, start=2):
            cell = ws.cell(row_idx, col_idx)
            val = df_requests.loc[emp, day]
            if (emp, day) in approved_set:
                if pd.notna(val) and val != 0:
                   cell.value = "公" if val == 2 else "休"
                   cell.fill = red_fill
                   cell.alignment = center_align
    wb.save(output)
    return output.getvalue()

col1, col2 = st.columns([1, 2])
with col1:
    st.subheader("步驟 1: 下載範本")
    st.download_button(
        label="📥 下載 Excel 輸入範本",
        data=generate_template_bytes(),
        file_name="排班輸入範本.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
with col2:
    st.subheader("步驟 2: 上傳並排班")
    uploaded_file = st.file_uploader("上傳填好的 Excel 檔案", type=["xlsx"])

if uploaded_file:
    st.divider()
    if st.button("🚀 開始排班", type="primary"):
        with st.spinner("運算中..."):
            raw_df = pd.read_excel(uploaded_file, sheet_name="Requests", index_col=0)
            result_df, approved_set, error_msg = process_schedule(uploaded_file, year, month, priority_emps)
            if error_msg:
                st.error(error_msg)
            else:
                st.success("✅ 完成！")
                st.subheader("📊 排班結果預覽")
                def highlight_approved(val):
                    color = '#ff9999' if val in ['休', '公'] else ''
                    return f'background-color: {color}'
                st.dataframe(result_df.style.map(highlight_approved), use_container_width=True, height=600)
                st.subheader("步驟 3: 下載結果")
                excel_data = create_download_excel(raw_df, approved_set, year, month)
                st.download_button(
                    label="📥 下載排班結果 Excel",
                    data=excel_data,
                    file_name=f"排班結果_{year}_{month}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
